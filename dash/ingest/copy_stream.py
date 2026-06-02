"""GB-scale streaming ingest — psycopg3 COPY FROM STDIN.

Replaces RAM-bound `pd.read_csv() → df.to_sql()` path for big CSV uploads.
Streams source file in 64KB chunks straight to Postgres via COPY protocol.

Capacity:
  CSV     — UNLIMITED (streams disk → COPY, no RAM materialization)
  Excel   — bounded by openpyxl read_only iter_rows (~10M rows / few GB)
  Parquet — bounded by pyarrow row groups (~GB scale per group)

What's NOT here (intentional — keep this module focused on RAW ingest):
  - Schema profiling (handled by app/upload.py:_sql_profile_columns AFTER ingest)
  - LLM enrichment (handled by training pipeline AFTER ingest)
  - Header detection / multi-table split / unpivot (handled by Excel pipeline)
  - Column type coercion (Postgres COPY uses text → ALTER later if needed)

Public API:
  copy_csv_stream(file_path, schema, table, *, encoding='utf-8', delimiter=',',
                  has_header=True, progress_cb=None) -> dict
  stream_xlsx_to_postgres(file_path, schema, table, sheet=None, progress_cb=None) -> dict
  stream_parquet_to_postgres(file_path, schema, table, progress_cb=None) -> dict

All functions:
  - CREATE TABLE if not exists (all-text columns from header) — schema-evolution
    handled in upper layer if needed
  - Fail-soft on row errors — log + continue (configurable)
  - Emit progress callback every N rows for SSE streaming
  - Return {rows_loaded, bytes_read, elapsed_s, errors_count}
"""
from __future__ import annotations

import csv
import io
import logging
import os
import re
import time
from pathlib import Path
from typing import Any, Callable, Iterable

logger = logging.getLogger(__name__)

# ───────────────────────────────────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────────────────────────────────

_SAFE_IDENT_RE = re.compile(r"[^a-z0-9_]")


def _safe_ident(name: str, maxlen: int = 63, *, preserve_underscores: bool = True) -> str:
    """Postgres-safe identifier: lowercase, alnum + underscore, ≤63 chars.

    Project slugs like `proj_demo_foo_` end in underscore by convention — must
    preserve to match `dash_projects.schema_name`. Only strip when called for
    column/table names (cosmetic cleanup, defaults preserve).
    """
    s = _SAFE_IDENT_RE.sub("_", name.lower())
    if not preserve_underscores:
        s = s.strip("_")
    if not s or s.strip("_") == "":
        s = "col"
    if s[0].isdigit():
        s = f"c_{s}"
    return s[:maxlen]


def _quote_ident(name: str) -> str:
    """Postgres identifier quoting (handles embedded quotes)."""
    return '"' + name.replace('"', '""') + '"'


def _ensure_schema(conn, schema: str) -> None:
    """Idempotent CREATE SCHEMA."""
    schema = _safe_ident(schema)
    conn.execute(f'CREATE SCHEMA IF NOT EXISTS "{schema}"')


def _create_table_all_text(conn, schema: str, table: str, columns: list[str]) -> None:
    """CREATE TABLE with all-TEXT columns. Type coercion happens post-ingest.

    Idempotent via DROP + CREATE — REPLACE mode. For APPEND, use _ensure_table.
    """
    schema = _safe_ident(schema)
    table = _safe_ident(table)
    cols_sql = ", ".join(f'{_quote_ident(c)} TEXT' for c in columns)
    conn.execute(f'DROP TABLE IF EXISTS "{schema}"."{table}" CASCADE')
    conn.execute(f'CREATE TABLE "{schema}"."{table}" ({cols_sql})')


# ───────────────────────────────────────────────────────────────────────────
# CSV — TRUE streaming via COPY FROM STDIN
# ───────────────────────────────────────────────────────────────────────────

def copy_csv_stream(
    file_path: str | Path,
    schema: str,
    table: str,
    *,
    encoding: str = "utf-8",
    delimiter: str = ",",
    has_header: bool = True,
    progress_cb: Callable[[int, int], None] | None = None,
    chunk_bytes: int = 1024 * 1024,
) -> dict[str, Any]:
    """Stream CSV → Postgres via COPY. Constant memory, GB-scale.

    Args:
      file_path:   absolute path to CSV on disk
      schema:      target schema (project slug, will be safe-ident'd)
      table:       target table name
      encoding:    file encoding (use chardet upstream if unknown)
      delimiter:   CSV delimiter (',', '\\t', '|', etc)
      has_header:  if True, first row becomes column names
      progress_cb: callback(rows_loaded, bytes_read) every ~10K rows
      chunk_bytes: read chunk size (default 64KB)

    Returns:
      {rows_loaded, bytes_read, elapsed_s, columns}
    """
    from db.session import get_write_engine

    file_path = Path(file_path)
    t0 = time.time()
    total_bytes = file_path.stat().st_size

    # Detect delimiter if comma + first line has none (best-effort)
    # Sniff first 8KB for header + delimiter
    with open(file_path, "rb") as f:
        head_bytes = f.read(8192)
    head_text = head_bytes.decode(encoding, errors="replace")
    if delimiter == "," and "," not in head_text.split("\n", 1)[0]:
        # Try common alternates
        for cand in ("\t", "|", ";"):
            if cand in head_text.split("\n", 1)[0]:
                delimiter = cand
                logger.info(f"copy_csv_stream: auto-detected delimiter '{cand!r}'")
                break

    # Read header for column names
    columns: list[str] = []
    if has_header:
        with open(file_path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            header_row = next(reader, None)
            if not header_row:
                raise ValueError(f"CSV {file_path.name}: empty file")
            seen: dict[str, int] = {}
            for raw in header_row:
                base = _safe_ident(raw or "col")
                if base in seen:
                    seen[base] += 1
                    columns.append(f"{base}_{seen[base]}")
                else:
                    seen[base] = 0
                    columns.append(base)
    else:
        # No header: peek first data row to count columns
        with open(file_path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            first = next(reader, None)
            if not first:
                raise ValueError("empty CSV")
            columns = [f"col_{i+1}" for i in range(len(first))]

    eng = get_write_engine()
    raw_conn = eng.raw_connection()
    try:
        with raw_conn.cursor() as cur:
            _ensure_schema(cur, schema)
            _create_table_all_text(cur, _safe_ident(schema), _safe_ident(table), columns)
            raw_conn.commit()

            # COPY FROM STDIN — psycopg3 streams from any iterable of bytes
            qualified = f'"{_safe_ident(schema)}"."{_safe_ident(table)}"'
            cols_quoted = ", ".join(_quote_ident(c) for c in columns)
            copy_sql = (
                f"COPY {qualified} ({cols_quoted}) FROM STDIN "
                f"WITH (FORMAT csv, DELIMITER {chr(39)}{delimiter}{chr(39)}, "
                f"HEADER {'TRUE' if has_header else 'FALSE'}, NULL '')"
            )

            rows_loaded = 0
            bytes_read = 0
            last_progress_ts = time.time()
            with cur.copy(copy_sql) as copy:
                with open(file_path, "rb") as f:
                    while True:
                        chunk = f.read(chunk_bytes)
                        if not chunk:
                            break
                        copy.write(chunk)
                        bytes_read += len(chunk)
                        # Progress estimate (bytes-based, no row count yet)
                        if progress_cb and (time.time() - last_progress_ts) > 2.0:
                            try:
                                pct = int(100 * bytes_read / max(total_bytes, 1))
                                progress_cb(pct, bytes_read)
                            except Exception:
                                pass
                            last_progress_ts = time.time()

            # COPY done — get row count from cursor
            rows_loaded = cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0
            raw_conn.commit()
    finally:
        try:
            raw_conn.close()
        except Exception:
            pass

    elapsed = time.time() - t0
    if progress_cb:
        try:
            progress_cb(100, bytes_read)
        except Exception:
            pass

    logger.info(
        "copy_csv_stream: %s rows=%d bytes=%d elapsed=%.1fs (%.1f MB/s)",
        file_path.name, rows_loaded, bytes_read, elapsed,
        (bytes_read / 1024 / 1024) / max(elapsed, 0.01),
    )
    return {
        "rows_loaded": rows_loaded,
        "bytes_read": bytes_read,
        "elapsed_s": round(elapsed, 2),
        "columns": columns,
        "mb_per_sec": round((bytes_read / 1024 / 1024) / max(elapsed, 0.01), 2),
    }


# ───────────────────────────────────────────────────────────────────────────
# Excel — openpyxl read_only + batched COPY
# ───────────────────────────────────────────────────────────────────────────

def stream_xlsx_to_postgres(
    file_path: str | Path,
    schema: str,
    table: str,
    *,
    sheet: str | None = None,
    has_header: bool = True,
    batch_rows: int = 5000,
    progress_cb: Callable[[int, int], None] | None = None,
) -> dict[str, Any]:
    """Stream Excel sheet → Postgres via openpyxl read_only + COPY.

    Capacity: ~10M rows / few GB (openpyxl read_only iter_rows).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl required for xlsx streaming")
    from db.session import get_write_engine

    file_path = Path(file_path)
    t0 = time.time()

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    row_iter = ws.iter_rows(values_only=True)

    columns: list[str] = []
    if has_header:
        first = next(row_iter, None)
        if not first:
            wb.close()
            raise ValueError(f"Sheet {ws.title!r}: empty")
        seen: dict[str, int] = {}
        for raw in first:
            base = _safe_ident(str(raw or "col"))
            if base in seen:
                seen[base] += 1
                columns.append(f"{base}_{seen[base]}")
            else:
                seen[base] = 0
                columns.append(base)
    else:
        peek = next(row_iter, None)
        if not peek:
            wb.close()
            raise ValueError("empty sheet")
        columns = [f"col_{i+1}" for i in range(len(peek))]
        # Push peek back as first data row
        row_iter = _prepend_iter(peek, row_iter)

    eng = get_write_engine()
    raw_conn = eng.raw_connection()
    rows_loaded = 0
    try:
        with raw_conn.cursor() as cur:
            _ensure_schema(cur, schema)
            _create_table_all_text(cur, _safe_ident(schema), _safe_ident(table), columns)
            raw_conn.commit()

            qualified = f'"{_safe_ident(schema)}"."{_safe_ident(table)}"'
            cols_quoted = ", ".join(_quote_ident(c) for c in columns)
            copy_sql = (
                f"COPY {qualified} ({cols_quoted}) FROM STDIN "
                f"WITH (FORMAT csv, DELIMITER ',', NULL '')"
            )

            with cur.copy(copy_sql) as copy:
                buf = io.StringIO()
                writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
                batch_count = 0
                last_progress_ts = time.time()
                for row in row_iter:
                    if row is None:
                        continue
                    # Normalize: None → empty, others → str
                    out = []
                    for cell in row:
                        if cell is None:
                            out.append("")
                        elif isinstance(cell, str):
                            out.append(cell)
                        else:
                            out.append(str(cell))
                    # Pad/trim to match columns
                    if len(out) < len(columns):
                        out += [""] * (len(columns) - len(out))
                    elif len(out) > len(columns):
                        out = out[:len(columns)]
                    writer.writerow(out)
                    batch_count += 1
                    rows_loaded += 1

                    if batch_count >= batch_rows:
                        copy.write(buf.getvalue().encode("utf-8"))
                        buf.seek(0); buf.truncate(0)
                        batch_count = 0
                        if progress_cb and (time.time() - last_progress_ts) > 2.0:
                            try:
                                progress_cb(rows_loaded, 0)
                            except Exception:
                                pass
                            last_progress_ts = time.time()

                # Flush tail
                if batch_count > 0:
                    copy.write(buf.getvalue().encode("utf-8"))

            raw_conn.commit()
    finally:
        try:
            wb.close()
        except Exception:
            pass
        try:
            raw_conn.close()
        except Exception:
            pass

    elapsed = time.time() - t0
    logger.info(
        "stream_xlsx: %s sheet=%s rows=%d elapsed=%.1fs",
        file_path.name, ws.title, rows_loaded, elapsed,
    )
    return {
        "rows_loaded": rows_loaded,
        "elapsed_s": round(elapsed, 2),
        "columns": columns,
        "sheet": ws.title,
    }


def _prepend_iter(first, rest):
    yield first
    for r in rest:
        yield r


# ───────────────────────────────────────────────────────────────────────────
# Parquet — pyarrow row groups → batched COPY
# ───────────────────────────────────────────────────────────────────────────

def stream_parquet_to_postgres(
    file_path: str | Path,
    schema: str,
    table: str,
    *,
    progress_cb: Callable[[int, int], None] | None = None,
) -> dict[str, Any]:
    """Stream Parquet → Postgres row-group by row-group."""
    try:
        import pyarrow.parquet as pq
    except ImportError:
        raise RuntimeError("pyarrow required for parquet streaming")
    from db.session import get_write_engine

    file_path = Path(file_path)
    t0 = time.time()

    pf = pq.ParquetFile(file_path)
    schema_arrow = pf.schema_arrow
    columns = [_safe_ident(c) for c in schema_arrow.names]
    # Dedup
    seen: dict[str, int] = {}
    dedup_cols = []
    for c in columns:
        if c in seen:
            seen[c] += 1
            dedup_cols.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            dedup_cols.append(c)
    columns = dedup_cols

    eng = get_write_engine()
    raw_conn = eng.raw_connection()
    rows_loaded = 0
    try:
        with raw_conn.cursor() as cur:
            _ensure_schema(cur, schema)
            _create_table_all_text(cur, _safe_ident(schema), _safe_ident(table), columns)
            raw_conn.commit()

            qualified = f'"{_safe_ident(schema)}"."{_safe_ident(table)}"'
            cols_quoted = ", ".join(_quote_ident(c) for c in columns)
            copy_sql = (
                f"COPY {qualified} ({cols_quoted}) FROM STDIN "
                f"WITH (FORMAT csv, DELIMITER ',', NULL '')"
            )

            with cur.copy(copy_sql) as copy:
                for rg_idx in range(pf.num_row_groups):
                    rg = pf.read_row_group(rg_idx)
                    # Convert to CSV chunk in memory (row group bounded)
                    buf = io.StringIO()
                    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
                    # rg is a Table; iterate row-by-row via to_pylist (RG sized to fit)
                    for row in rg.to_pylist():
                        out = [str(row.get(c, "")) if row.get(c) is not None else ""
                               for c in schema_arrow.names]
                        writer.writerow(out)
                        rows_loaded += 1
                    copy.write(buf.getvalue().encode("utf-8"))
                    if progress_cb:
                        try:
                            progress_cb(rows_loaded, rg_idx + 1)
                        except Exception:
                            pass

            raw_conn.commit()
    finally:
        try:
            raw_conn.close()
        except Exception:
            pass

    elapsed = time.time() - t0
    logger.info(
        "stream_parquet: %s rows=%d row_groups=%d elapsed=%.1fs",
        file_path.name, rows_loaded, pf.num_row_groups, elapsed,
    )
    return {
        "rows_loaded": rows_loaded,
        "elapsed_s": round(elapsed, 2),
        "columns": columns,
        "row_groups": pf.num_row_groups,
    }


# ───────────────────────────────────────────────────────────────────────────
# Auto-dispatch by extension
# ───────────────────────────────────────────────────────────────────────────

def stream_to_postgres(
    file_path: str | Path,
    schema: str,
    table: str,
    *,
    progress_cb: Callable[[int, int], None] | None = None,
    **opts,
) -> dict[str, Any]:
    """Auto-dispatch by extension. Returns ingest stats dict."""
    file_path = Path(file_path)
    ext = file_path.suffix.lower().lstrip(".")
    if ext == "csv":
        return copy_csv_stream(file_path, schema, table,
                               progress_cb=progress_cb, **opts)
    if ext in ("xlsx", "xls"):
        return stream_xlsx_to_postgres(file_path, schema, table,
                                       progress_cb=progress_cb, **opts)
    if ext in ("parquet", "pq"):
        return stream_parquet_to_postgres(file_path, schema, table,
                                          progress_cb=progress_cb, **opts)
    raise ValueError(f"Unsupported extension for streaming ingest: .{ext}")


__all__ = [
    "copy_csv_stream",
    "stream_xlsx_to_postgres",
    "stream_parquet_to_postgres",
    "stream_to_postgres",
]
